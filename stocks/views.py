"""
Stocks Views
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db import models
from django.db.models import Q, F, Sum
from django.core.paginator import Paginator
from decimal import Decimal, InvalidOperation

from .models import Product, ProductCategory, UnitOfMeasure, StockMovement, StockBatch, ProductPackaging
from .forms import ProductForm, CategoryForm, UnitForm, StockAdjustmentForm, ProductPackagingForm
from .services import StockManagementService, BarcodeService
from decorators.decorators import group_required


def _sync_packaging_prices(product, saved_packaging):
    """
    When a packaging is saved, sync related packaging prices and the product's base prices.
    - If bulk saved: derive display and unit prices, update product.sale_price/purchase_price
    - If display saved: update unit prices and product
    - If unit saved: update product
    """
    unit_pkg = ProductPackaging.objects.filter(product=product, packaging_type='unit', is_active=True).first()
    display_pkg = ProductPackaging.objects.filter(product=product, packaging_type='display', is_active=True).first()
    bulk_pkg = ProductPackaging.objects.filter(product=product, packaging_type='bulk', is_active=True).first()
    
    # Derive unit prices from the highest-level packaging available
    if bulk_pkg:
        total_units = bulk_pkg.units_quantity or 1
        unit_purchase = bulk_pkg.purchase_price / total_units if total_units > 0 else Decimal('0')
        unit_sale = bulk_pkg.sale_price / total_units if total_units > 0 else Decimal('0')
        
        # Update display packaging derived prices if it doesn't have its own sale price
        if display_pkg and saved_packaging.packaging_type == 'bulk':
            displays_per_bulk = bulk_pkg.displays_per_bulk or 1
            if display_pkg.purchase_price == Decimal('0'):
                display_pkg.purchase_price = bulk_pkg.purchase_price / displays_per_bulk
            if display_pkg.sale_price == Decimal('0'):
                display_pkg.sale_price = bulk_pkg.sale_price / displays_per_bulk
                if display_pkg.purchase_price > 0:
                    display_pkg.margin_percent = ((display_pkg.sale_price - display_pkg.purchase_price) / display_pkg.purchase_price) * 100
                display_pkg.save()
        
        # Update unit packaging derived prices if it doesn't have its own sale price
        if unit_pkg and saved_packaging.packaging_type in ('bulk', 'display'):
            if unit_pkg.purchase_price == Decimal('0'):
                unit_pkg.purchase_price = unit_purchase
            if unit_pkg.sale_price == Decimal('0'):
                unit_pkg.sale_price = unit_sale
                if unit_pkg.purchase_price > 0:
                    unit_pkg.margin_percent = ((unit_pkg.sale_price - unit_pkg.purchase_price) / unit_pkg.purchase_price) * 100
                unit_pkg.save()
        
        # Always update the Product's base prices from unit-level
        product.purchase_price = unit_purchase
        product.sale_price = unit_sale
        product.save(update_fields=['purchase_price', 'sale_price'])
    
    elif display_pkg:
        units_per_display = display_pkg.units_per_display or 1
        unit_purchase = display_pkg.purchase_price / units_per_display
        unit_sale = display_pkg.sale_price / units_per_display
        
        if unit_pkg and saved_packaging.packaging_type == 'display':
            if unit_pkg.purchase_price == Decimal('0'):
                unit_pkg.purchase_price = unit_purchase
            if unit_pkg.sale_price == Decimal('0'):
                unit_pkg.sale_price = unit_sale
                if unit_pkg.purchase_price > 0:
                    unit_pkg.margin_percent = ((unit_pkg.sale_price - unit_pkg.purchase_price) / unit_pkg.purchase_price) * 100
                unit_pkg.save()
        
        product.purchase_price = unit_purchase
        product.sale_price = unit_sale
        product.save(update_fields=['purchase_price', 'sale_price'])
    
    elif unit_pkg:
        product.purchase_price = unit_pkg.purchase_price
        product.sale_price = unit_pkg.sale_price
        product.save(update_fields=['purchase_price', 'sale_price'])


@login_required
@group_required(['Admin', 'Cajero Manager', 'Cashier'])
def product_list(request):
    """List all products."""
    products = Product.objects.select_related('category', 'unit_of_measure')
    
    # Filters
    search = request.GET.get('search', '')
    category = request.GET.get('category', '')
    status = request.GET.get('status', '')
    stock_alert = request.GET.get('stock_alert', '')
    
    if search:
        if search.isdigit():
            products = products.filter(
                Q(sku__istartswith=search) |
                Q(barcode__istartswith=search)
            )
        else:
            products = products.filter(
                Q(name__icontains=search) |
                Q(sku__icontains=search) |
                Q(barcode__icontains=search)
            )
    
    if category:
        products = products.filter(category_id=category)
    
    if status == 'active':
        products = products.filter(is_active=True)
    elif status == 'inactive':
        products = products.filter(is_active=False)
    
    if stock_alert == 'low':
        products = products.filter(current_stock__lte=F('min_stock'))
    elif stock_alert == 'out':
        products = products.filter(current_stock=0)
    
    # Pagination
    paginator = Paginator(products, 20)
    page = request.GET.get('page', 1)
    products = paginator.get_page(page)
    
    categories = ProductCategory.objects.filter(is_active=True)
    
    context = {
        'products': products,
        'categories': categories,
        'search': search,
        'selected_category': category,
        'selected_status': status,
        'stock_alert': stock_alert,
    }
    
    return render(request, 'stocks/product_list.html', context)


def _save_inline_packaging(request, product):
    """Process inline packaging fields from the product form."""
    units_per_display = int(request.POST.get('pkg_units_per_display', 1) or 1)
    displays_per_bulk = int(request.POST.get('pkg_displays_per_bulk', 1) or 1)
    total_units = units_per_display * displays_per_bulk

    def _calc_margin(purchase, sale):
        if purchase and purchase > 0 and sale and sale > 0:
            return ((sale - purchase) / purchase) * 100
        return Decimal('0')

    def _check_barcode(barcode, pkg_type):
        """Validate barcode is not used by another packaging record."""
        if not barcode:
            return None
        dup = ProductPackaging.objects.filter(barcode=barcode).exclude(
            product=product, packaging_type=pkg_type
        ).first()
        if dup:
            raise ValueError(
                f'El código de barras {barcode} ya está en uso por '
                f'{dup.product.name} - {dup.get_packaging_type_display()}'
            )
        return barcode

    # Get bulk purchase price to derive display/unit costs
    b_purchase = Decimal(request.POST.get('bulk_purchase_price', '0').strip() or '0')

    # Bulk packaging
    if request.POST.get('has_bulk'):
        b_barcode = _check_barcode(request.POST.get('bulk_barcode', '').strip(), 'bulk')
        b_name = request.POST.get('bulk_name', '').strip()
        b_sale = Decimal(request.POST.get('bulk_sale_price', '0').strip() or '0')

        ProductPackaging.objects.update_or_create(
            product=product, packaging_type='bulk',
            defaults={
                'barcode': b_barcode or None,
                'name': b_name or f'Bulto x {total_units}',
                'units_per_display': units_per_display,
                'displays_per_bulk': displays_per_bulk,
                'purchase_price': b_purchase,
                'sale_price': b_sale,
                'margin_percent': _calc_margin(b_purchase, b_sale),
                'is_active': True,
            }
        )

    # Display packaging — purchase price = bulk / displays_per_bulk
    if request.POST.get('has_display'):
        d_barcode = _check_barcode(request.POST.get('display_barcode', '').strip(), 'display')
        d_name = request.POST.get('display_name', '').strip()
        d_purchase = (b_purchase / displays_per_bulk) if b_purchase > 0 else Decimal('0')
        d_sale = Decimal(request.POST.get('display_sale_price', '0').strip() or '0')

        ProductPackaging.objects.update_or_create(
            product=product, packaging_type='display',
            defaults={
                'barcode': d_barcode or None,
                'name': d_name or f'Display x {units_per_display}',
                'units_per_display': units_per_display,
                'displays_per_bulk': 1,
                'purchase_price': d_purchase,
                'sale_price': d_sale,
                'margin_percent': _calc_margin(d_purchase, d_sale),
                'is_active': True,
            }
        )

    # Unit packaging — purchase price = bulk / total_units
    if request.POST.get('has_unit'):
        u_barcode = _check_barcode(request.POST.get('unit_barcode', '').strip(), 'unit')
        u_name = request.POST.get('unit_name', '').strip()
        u_purchase = (b_purchase / total_units) if b_purchase > 0 else Decimal('0')
        u_sale = Decimal(request.POST.get('unit_sale_price', '0').strip() or '0')

        ProductPackaging.objects.update_or_create(
            product=product, packaging_type='unit',
            defaults={
                'barcode': u_barcode or None,
                'name': u_name or 'Unidad',
                'units_per_display': 1,
                'displays_per_bulk': 1,
                'purchase_price': u_purchase,
                'sale_price': u_sale,
                'margin_percent': _calc_margin(u_purchase, u_sale),
                'is_active': True,
            }
        )


@login_required
@group_required(['Admin', 'Cajero Manager'])
def product_create(request):
    """Create new product."""
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            product = form.save()
            # Process inline packaging if provided
            _save_inline_packaging(request, product)
            messages.success(request, f'Producto "{product.name}" creado correctamente.')
            return redirect('stocks:product_list')
    else:
        form = ProductForm()
    
    return render(request, 'stocks/product_form.html', {
        'form': form,
        'title': 'Nuevo Producto'
    })


@login_required
@group_required(['Admin', 'Cajero Manager'])
def product_edit(request, pk):
    """Edit product."""
    product = get_object_or_404(Product, pk=pk)

    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES, instance=product)
        if form.is_valid():
            form.save()
            messages.success(request, f'Producto "{product.name}" actualizado correctamente.')
            return redirect('stocks:product_list')
    else:
        form = ProductForm(instance=product)

    return render(request, 'stocks/product_form.html', {
        'form': form,
        'title': 'Editar Producto',
        'product': product,
    })


@login_required
@group_required(['Admin', 'Cajero Manager'])
def product_delete(request, pk):
    """Delete product (soft delete)."""
    product = get_object_or_404(Product, pk=pk)
    
    if request.method == 'POST':
        product.is_active = False
        product.save()
        messages.success(request, f'Producto "{product.name}" desactivado correctamente.')
        return redirect('stocks:product_list')
    
    return render(request, 'stocks/product_confirm_delete.html', {'product': product})


@login_required
@group_required(['Admin', 'Cajero Manager'])
def product_detail(request, pk):
    """Product detail view."""
    product = get_object_or_404(Product, pk=pk)
    movements = product.stock_movements.order_by('-created_at')[:20]
    active_batches = product.batches.filter(quantity_remaining__gt=0).order_by('-purchased_at')[:5]

    return render(request, 'stocks/product_detail.html', {
        'product': product,
        'movements': movements,
        'active_batches': active_batches,
    })


@login_required
@group_required(['Admin', 'Cajero Manager'])
def stock_adjust(request, pk):
    """Adjust product stock."""
    product = get_object_or_404(Product, pk=pk)
    
    if request.method == 'POST':
        new_quantity = request.POST.get('new_quantity')
        reason = request.POST.get('reason', '')
        notes = request.POST.get('notes', '')
        batch_purchase_price = request.POST.get('purchase_price', '')
        batch_supplier = request.POST.get('supplier_name', '')

        # Mapear motivos a texto legible
        reason_map = {
            'conteo_fisico': 'Conteo Físico / Inventario',
            'mercaderia_danada': 'Mercadería Dañada',
            'mercaderia_vencida': 'Mercadería Vencida',
            'robo_perdida': 'Robo / Pérdida',
            'devolucion': 'Devolución',
            'correccion_error': 'Corrección de Error',
            'consumo_interno': 'Consumo Interno',
            'otro': 'Otro',
        }

        reason_text = reason_map.get(reason, reason)
        if notes:
            reason_text = f"{reason_text}: {notes}"

        try:
            from decimal import Decimal
            from django.db import transaction
            from django.utils import timezone as tz

            new_quantity = Decimal(new_quantity)
            old_quantity = product.current_stock
            diff = new_quantity - old_quantity

            with transaction.atomic():
                StockManagementService.adjust_stock(
                    product=product,
                    new_quantity=new_quantity,
                    reason=reason_text,
                    user=request.user
                )

                # Create StockBatch for entry adjustments (stock increase)
                if diff > 0 and batch_purchase_price:
                    StockBatch.objects.create(
                        product=product,
                        supplier_name=batch_supplier or '',
                        quantity_purchased=diff,
                        quantity_remaining=diff,
                        purchase_price=Decimal(batch_purchase_price),
                        purchased_at=tz.now(),
                        created_by=request.user,
                        notes=f'Ajuste de stock: {reason_text}',
                    )

            messages.success(request, f'Stock de "{product.name}" ajustado correctamente.')
            return redirect('stocks:product_detail', pk=pk)
        except Exception as e:
            messages.error(request, f'Error al ajustar stock: {str(e)}')
    
    form = StockAdjustmentForm(initial={'new_quantity': product.current_stock})
    
    return render(request, 'stocks/stock_adjust.html', {
        'form': form,
        'product': product
    })


@login_required
@group_required(['Admin', 'Cajero Manager'])
def product_movement_list(request, pk=None):
    """
    Kardex: historial de movimientos de stock.
    Si pk tiene valor, filtra por producto. Si no, muestra todos.
    """
    product = None
    movements = StockMovement.objects.select_related('product', 'created_by').all()

    if pk:
        product = get_object_or_404(Product, pk=pk)
        movements = movements.filter(product=product)

    # Filters
    search = request.GET.get('search', '')
    movement_type = request.GET.get('type', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    if search and not pk:
        movements = movements.filter(
            Q(product__name__icontains=search) |
            Q(product__sku__icontains=search) |
            Q(product__barcode__icontains=search)
        )

    if movement_type:
        movements = movements.filter(movement_type=movement_type)

    if date_from:
        movements = movements.filter(created_at__date__gte=date_from)

    if date_to:
        movements = movements.filter(created_at__date__lte=date_to)

    paginator = Paginator(movements, 50)
    page = request.GET.get('page', 1)
    movements_page = paginator.get_page(page)

    return render(request, 'stocks/movement_list.html', {
        'movements': movements_page,
        'product': product,
        'search': search,
        'selected_type': movement_type,
        'date_from': date_from,
        'date_to': date_to,
        'movement_types': StockMovement.MOVEMENT_TYPES,
    })


@login_required
@group_required(['Admin', 'Cajero Manager'])
def cost_history(request, pk=None):
    """
    Historial de costos de compra — tracking de márgenes por proveedor.
    Si pk tiene valor, filtra por producto. Si no, muestra todos.
    """
    from django.db.models import Sum, Avg, Min, Max, Count

    product = None
    batches = StockBatch.objects.select_related('product', 'created_by', 'purchase').all()

    if pk:
        product = get_object_or_404(Product, pk=pk)
        batches = batches.filter(product=product)

    # Filters
    search = request.GET.get('search', '')
    supplier = request.GET.get('supplier', '')
    show_depleted = request.GET.get('depleted', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    if search and not pk:
        batches = batches.filter(
            Q(product__name__icontains=search) |
            Q(product__sku__icontains=search) |
            Q(supplier_name__icontains=search)
        )

    if supplier:
        batches = batches.filter(supplier_name__icontains=supplier)

    if not show_depleted:
        batches = batches.filter(quantity_remaining__gt=0)

    if date_from:
        batches = batches.filter(purchased_at__date__gte=date_from)

    if date_to:
        batches = batches.filter(purchased_at__date__lte=date_to)

    # Summary metrics (only active stock)
    active_qs = batches.filter(quantity_remaining__gt=0)
    agg = active_qs.aggregate(
        total_invested=Sum(
            F('quantity_purchased') * F('purchase_price'),
            output_field=models.DecimalField()
        ),
        total_remaining_cost=Sum(
            F('quantity_remaining') * F('purchase_price'),
            output_field=models.DecimalField()
        ),
        total_remaining_qty=Sum('quantity_remaining'),
        avg_purchase_price=Avg('purchase_price'),
        min_purchase_price=Min('purchase_price'),
        max_purchase_price=Max('purchase_price'),
        entry_count=Count('id'),
    )
    total_invested = agg['total_invested'] or Decimal('0')
    total_remaining_cost = agg['total_remaining_cost'] or Decimal('0')
    total_remaining_qty = agg['total_remaining_qty'] or Decimal('0')
    avg_purchase_price = agg['avg_purchase_price'] or Decimal('0')
    min_purchase_price = agg['min_purchase_price'] or Decimal('0')
    max_purchase_price = agg['max_purchase_price'] or Decimal('0')
    entry_count = agg['entry_count'] or 0

    # Weighted average purchase price (more accurate than simple avg)
    if total_remaining_qty > 0:
        weighted_avg_cost = (total_remaining_cost / total_remaining_qty).quantize(Decimal('0.01'))
    else:
        weighted_avg_cost = Decimal('0')

    # Potential revenue if all remaining stock sold at current list price
    potential_revenue = Decimal('0')
    if product and product.sale_price:
        potential_revenue = total_remaining_qty * product.sale_price
    potential_profit = potential_revenue - total_remaining_cost

    # Per-supplier breakdown (for the filtered product if pk, else all)
    supplier_stats = (
        active_qs
        .values('supplier_name')
        .annotate(
            qty=Sum('quantity_remaining'),
            total_cost=Sum(F('quantity_remaining') * F('purchase_price'), output_field=models.DecimalField()),
            avg_cost=Avg('purchase_price'),
            min_cost=Min('purchase_price'),
            max_cost=Max('purchase_price'),
            entries=Count('id'),
        )
        .order_by('-total_cost')
    )

    # Available supplier names for filter dropdown
    all_suppliers = (
        StockBatch.objects
        .exclude(supplier_name='')
        .values_list('supplier_name', flat=True)
        .distinct()
        .order_by('supplier_name')
    )

    paginator = Paginator(batches, 50)
    page = request.GET.get('page', 1)
    batches_page = paginator.get_page(page)

    return render(request, 'stocks/cost_history.html', {
        'batches': batches_page,
        'product': product,
        'search': search,
        'supplier': supplier,
        'show_depleted': show_depleted,
        'date_from': date_from,
        'date_to': date_to,
        'total_invested': total_invested,
        'total_remaining_cost': total_remaining_cost,
        'total_remaining_qty': total_remaining_qty,
        'weighted_avg_cost': weighted_avg_cost,
        'min_purchase_price': min_purchase_price,
        'max_purchase_price': max_purchase_price,
        'entry_count': entry_count,
        'potential_revenue': potential_revenue,
        'potential_profit': potential_profit,
        'supplier_stats': supplier_stats,
        'all_suppliers': all_suppliers,
    })


@login_required
@group_required(['Admin', 'Cajero Manager'])
def category_list(request):
    """List categories."""
    categories = ProductCategory.objects.all()
    return render(request, 'stocks/category_list.html', {'categories': categories})


@login_required
@group_required(['Admin', 'Cajero Manager'])
def category_create(request):
    """Create category."""
    if request.method == 'POST':
        form = CategoryForm(request.POST)
        if form.is_valid():
            category = form.save()
            messages.success(request, f'Categoría "{category.name}" creada correctamente.')
            return redirect('stocks:category_list')
    else:
        form = CategoryForm()
    
    return render(request, 'stocks/category_form.html', {
        'form': form,
        'title': 'Nueva Categoría'
    })


@login_required
@group_required(['Admin', 'Cajero Manager'])
def category_edit(request, pk):
    """Edit category."""
    category = get_object_or_404(ProductCategory, pk=pk)
    
    if request.method == 'POST':
        form = CategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            messages.success(request, f'Categoría "{category.name}" actualizada correctamente.')
            return redirect('stocks:category_list')
    else:
        form = CategoryForm(instance=category)
    
    return render(request, 'stocks/category_form.html', {
        'form': form,
        'title': 'Editar Categoría',
        'category': category
    })


@login_required
@group_required(['Admin', 'Cajero Manager'])
def low_stock_products(request):
    """List products with low stock."""
    products = StockManagementService.get_low_stock_products()
    return render(request, 'stocks/low_stock.html', {'products': products})


@login_required
def price_list(request):
    """Price list view."""
    products = Product.objects.filter(is_active=True).select_related('category')
    
    category = request.GET.get('category', '')
    if category:
        products = products.filter(category_id=category)
    
    categories = ProductCategory.objects.filter(is_active=True)
    
    return render(request, 'stocks/price_list.html', {
        'products': products,
        'categories': categories,
        'selected_category': category
    })


# API Endpoints

@login_required
def api_search_products(request):
    """API: Search products."""
    query = request.GET.get('q', '')
    
    if not query or len(query) < 2:
        return JsonResponse({'products': []})
    
    products = Product.objects.filter(is_active=True)
    
    # Check if it's a barcode search (8-13 digits) - exact match
    if query.isdigit() and 8 <= len(query) <= 13:
        products = products.filter(barcode=query)
    elif query.isdigit():
        products = products.filter(
            Q(sku__istartswith=query) |
            Q(barcode__istartswith=query)
        )
    else:
        products = products.filter(
            Q(name__icontains=query) |
            Q(sku__icontains=query) |
            Q(barcode__icontains=query)
        )
    
    products = products[:20]
    
    data = {
        'products': [
            {
                'id': p.id,
                'name': p.name,
                'sku': p.sku,
                'barcode': p.barcode or '',
                'sale_price': float(p.sale_price),
                'current_stock': float(p.current_stock),
                'unit': p.unit_of_measure.abbreviation if p.unit_of_measure else 'u'
            }
            for p in products
        ]
    }
    
    return JsonResponse(data)


@login_required
def api_generate_barcode(request):
    """API: Generate new barcode."""
    barcode = BarcodeService.generate_ean13()
    
    # Ensure it's unique
    while Product.objects.filter(barcode=barcode).exists():
        barcode = BarcodeService.generate_ean13()
    
    return JsonResponse({'barcode': barcode})


# ==================== IMPORTAR EXCEL ====================

@login_required
@group_required(['Admin', 'Cajero Manager'])
def import_excel(request):
    """Importar productos desde Excel. Cada hoja = una categoría."""
    import openpyxl
    from django.db import transaction

    if request.method == 'POST':

        # ── Paso 2: Confirmar importación (no necesita archivo) ──
        if 'confirm' in request.POST:
            data = request.session.pop('_import_excel_data', None)
            if not data:
                messages.error(request, 'Sesión expirada. Volvé a subir el archivo.')
                return redirect('stocks:import_excel')

            flush = request.POST.get('flush') == '1'
            created = 0
            updated = 0
            cat_created = 0
            errors = []

            with transaction.atomic():
                # Si flush, borrar todo el inventario, promociones y categorías
                if flush:
                    from promotions.models import Promotion
                    from pos.models import QuickAccessButton, POSTransactionItem
                    from purchase.models import PurchaseItem
                    from sales.models import SaleItem
                    # Borrar items que referencian productos (PROTECT)
                    POSTransactionItem.objects.all().delete()
                    PurchaseItem.objects.all().delete()
                    SaleItem.objects.all().delete()
                    # Borrar en cascada
                    QuickAccessButton.objects.all().delete()
                    Promotion.objects.all().delete()
                    StockMovement.objects.all().delete()
                    ProductPackaging.objects.all().delete()
                    Product.objects.all().delete()
                    ProductCategory.objects.all().delete()
                    UnitOfMeasure.objects.all().delete()

                for sheet_data in data:
                    cat_name = sheet_data['category_name'].strip()

                    # Buscar categoría existente (case-insensitive)
                    category = ProductCategory.objects.filter(name__iexact=cat_name).first()
                    if not category:
                        category = ProductCategory.objects.create(name=cat_name)
                        cat_created += 1

                    for item in sheet_data['items']:
                        try:
                            barcode_val = _clean_barcode(item.get('barcode'))
                            sku_val = str(item.get('sku', '')).strip() if item.get('sku') else ''

                            product = None
                            # Buscar por barcode primero, luego por SKU
                            if barcode_val:
                                product = Product.objects.filter(barcode=barcode_val).first()
                            if not product and sku_val:
                                product = Product.objects.filter(sku=sku_val).first()

                            purchase_price = Decimal(str(item['purchase_price'])) if item.get('purchase_price') else Decimal('0.00')
                            sale_price = Decimal(str(item['sale_price'])) if item.get('sale_price') else Decimal('0.01')

                            uom = None
                            if item.get('unit'):
                                uom = _get_or_create_unit(item['unit'])

                            if product:
                                # Actualizar existente
                                product.category = category
                                product.purchase_price = purchase_price
                                product.sale_price = sale_price
                                product.cost_price = purchase_price
                                if uom:
                                    product.unit_of_measure = uom
                                if barcode_val and not product.barcode:
                                    product.barcode = barcode_val
                                product.save()
                                updated += 1
                            else:
                                # Crear nuevo - generar SKU si no tiene
                                sku = sku_val or f"IMP-{Product.objects.count() + 1:06d}"
                                base_sku = sku
                                counter = 1
                                while Product.objects.filter(sku=sku).exists():
                                    sku = f"{base_sku}-{counter}"
                                    counter += 1

                                # Evitar barcode duplicado
                                if barcode_val and Product.objects.filter(barcode=barcode_val).exists():
                                    barcode_val = None

                                Product.objects.create(
                                    sku=sku,
                                    barcode=barcode_val if barcode_val else None,
                                    name=item['nombre'],
                                    category=category,
                                    unit_of_measure=uom,
                                    purchase_price=purchase_price,
                                    sale_price=sale_price,
                                    cost_price=purchase_price,
                                    is_active=True,
                                )
                                created += 1
                        except Exception as e:
                            errors.append(f"{item.get('nombre', '???')}: {e}")

            if flush:
                messages.info(request, 'Se borró todo el inventario anterior.')
            msg = f'Importación completada: {created} creados, {updated} actualizados'
            if cat_created:
                msg += f', {cat_created} categorías nuevas'
            if errors:
                msg += f'. {len(errors)} errores.'
                for err in errors[:10]:
                    messages.warning(request, err)
            messages.success(request, msg)
            return redirect('stocks:product_list')

        # ── Paso 1: Preview (necesita archivo) ──
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, 'Seleccioná un archivo Excel.')
            return redirect('stocks:import_excel')

        if not excel_file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'El archivo debe ser .xlsx o .xls')
            return redirect('stocks:import_excel')

        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
        except Exception as e:
            messages.error(request, f'No se pudo leer el archivo: {e}')
            return redirect('stocks:import_excel')

        preview_data = []
        _debug_headers = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                continue

            header = [str(c).strip().lower() if c else '' for c in rows[0]]
            col_map = _map_columns(header)

            # Si no detectó 'nombre', intentar con la primera columna de texto
            if 'nombre' not in col_map:
                for test_idx, h in enumerate(header):
                    if test_idx not in col_map.values() and h:
                        col_map['nombre'] = test_idx
                        break

            if 'nombre' not in col_map:
                _debug_headers.append(f"Hoja '{sheet_name}': columnas={header}, no se pudo detectar 'nombre'")
                continue

            items = []
            for row in rows[1:]:
                if not any(row):
                    continue
                item = _extract_row(row, col_map)
                if item['nombre']:
                    exists = False
                    bc = _clean_barcode(item.get('barcode'))
                    if bc:
                        exists = Product.objects.filter(barcode=bc).exists()
                        item['barcode'] = bc
                    if not exists and item.get('sku'):
                        exists = Product.objects.filter(sku=str(item['sku']).strip()).exists()
                    item['exists'] = exists
                    items.append(item)

            if items:
                preview_data.append({
                    'category_name': sheet_name.strip(),
                    'items': items,
                    'count': len(items),
                })

        if not preview_data:
            msg = 'No se encontraron datos válidos en el archivo.'
            if _debug_headers:
                for dbg in _debug_headers:
                    messages.info(request, dbg)
            else:
                sheet_info = [sn for sn in wb.sheetnames]
                messages.info(request, f'Hojas encontradas: {", ".join(sheet_info)}')
            messages.warning(request, msg)
            return redirect('stocks:import_excel')

        # Guardar datos en sesión para el paso de confirmación
        request.session['_import_excel_data'] = _serialize_preview(preview_data)
        return render(request, 'stocks/import_excel.html', {
            'preview': preview_data,
            'total_products': sum(s['count'] for s in preview_data),
            'total_categories': len(preview_data),
        })

    return render(request, 'stocks/import_excel.html')


def _map_columns(header):
    """Mapear nombres de columnas flexibles a campos internos."""
    import re
    col_map = {}

    # Patrones regex para cada campo - orden importa (más específico primero)
    patterns = [
        ('barcode', r'c[oó]d.*barra|barcode|ean|cod\.?\s*barra'),
        ('sku', r'c[oó]d.*interno|cod\.?\s*interno|sku|c[oó]digo(?!.*barra)|cod(?!.*barra)\b|interno'),
        ('nombre', r'nombre|producto|descripci[oó]n|art[ií]culo|detalle'),
        ('unit', r'unidad|u\.?m\.?|medida|uni\b|und\b'),
        ('margin', r'marg|markup|ganancia|rentab|%'),
        ('purchase_price', r'costo|compra|p\.?\s*costo|p\.?\s*compra'),
        ('sale_price', r'venta|p\.?\s*venta|pvp|precio(?!.*cost|.*compr)|publico|p[uú]blico'),
    ]

    for idx, col_name in enumerate(header):
        if not col_name:
            continue
        normalized = re.sub(r'[\s]+', ' ', col_name.strip().lower())
        for field, pattern in patterns:
            if field not in col_map and re.search(pattern, normalized):
                col_map[field] = idx
                break

    return col_map


def _extract_row(row, col_map):
    """Extraer datos de una fila usando el mapeo de columnas."""
    def get_val(field):
        idx = col_map.get(field)
        if idx is not None and idx < len(row):
            return row[idx]
        return None

    def to_str(val):
        if val is None:
            return ''
        return str(val).strip()

    def to_decimal(val):
        if val is None:
            return None
        if isinstance(val, (int, float)):
            return round(val, 2)
        s = str(val).strip().replace('$', '').strip()
        # Si tiene punto y coma, asumir formato argentino (1.234,56)
        if ',' in s and '.' in s:
            s = s.replace('.', '').replace(',', '.')
        elif ',' in s:
            # Solo coma: puede ser decimal (1234,56)
            s = s.replace(',', '.')
        # Si solo tiene punto, dejarlo como está (1234.56)
        try:
            return round(float(s), 2)
        except (ValueError, TypeError):
            return None

    nombre = to_str(get_val('nombre'))
    barcode = _clean_barcode(get_val('barcode'))
    sku_raw = get_val('sku')
    # SKU puede venir como número float en Excel
    sku = ''
    if sku_raw is not None:
        s = str(sku_raw).strip()
        if s.endswith('.0'):
            s = s[:-2]
        sku = s
    unit = to_str(get_val('unit'))
    margin = to_decimal(get_val('margin'))
    purchase_price = to_decimal(get_val('purchase_price'))
    sale_price = to_decimal(get_val('sale_price'))

    # Si hay margen y precio de costo pero no de venta, calcular
    if purchase_price and margin and not sale_price:
        sale_price = round(purchase_price * (1 + margin / 100), 2)

    # Si hay margen y precio de venta pero no de costo, calcular inverso
    if sale_price and margin and not purchase_price:
        purchase_price = round(sale_price / (1 + margin / 100), 2)

    return {
        'nombre': nombre,
        'barcode': barcode,
        'sku': sku,
        'unit': unit,
        'margin': margin,
        'purchase_price': purchase_price,
        'sale_price': sale_price,
    }


def _get_or_create_unit(name):
    """Buscar o crear unidad de medida."""
    if not name:
        return None
    name_lower = name.strip().lower()
    # Buscar por nombre o abreviatura
    uom = UnitOfMeasure.objects.filter(
        Q(name__iexact=name_lower) | Q(abbreviation__iexact=name_lower)
    ).first()
    if not uom:
        uom = UnitOfMeasure.objects.create(
            name=name.strip().title(),
            abbreviation=name.strip()[:10].upper(),
        )
    return uom


def _serialize_preview(preview_data):
    """Serializar preview para guardar en sesión."""
    result = []
    for sheet in preview_data:
        result.append({
            'category_name': sheet['category_name'],
            'items': sheet['items'],
        })
    return result


def _clean_barcode(val):
    """Limpiar código de barras: quitar .0 de floats, espacios, etc."""
    if not val:
        return ''
    s = str(val).strip()
    # Excel guarda números como float: 7790001234567.0 → 7790001234567
    if s.endswith('.0'):
        s = s[:-2]
    # Quitar cualquier punto o espacio
    s = s.replace('.', '').replace(' ', '')
    # Solo dejar dígitos
    if s and not s.isdigit():
        # Intentar extraer solo dígitos
        import re
        digits = re.sub(r'[^\d]', '', s)
        return digits if digits else ''
    return s


@login_required
def export_products_excel(request):
    """Export products to Excel — 4 sheets: Inventario, Stock Bajo, Márgenes, Valor."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.utils import timezone as dj_tz

    products = (
        Product.objects.filter(is_active=True)
        .select_related('category', 'unit_of_measure')
        .order_by('category__name', 'name')
    )

    # --- style helpers ---
    C_PURPLE = '2D1E5F'
    C_PINK   = 'E91E8C'
    C_LGRAY  = 'F2F2F2'
    C_WHITE  = 'FFFFFF'
    C_RED    = 'FFCCCC'
    C_ORANGE = 'FFE5CC'
    C_GREEN  = 'CCFFDD'

    def fill(c):
        return PatternFill(start_color=c, end_color=c, fill_type='solid')

    def border():
        s = Side(style='thin', color='CCCCCC')
        return Border(left=s, right=s, top=s, bottom=s)

    def header_row(ws, row_num, headers, color=C_PURPLE):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col, value=h)
            cell.font = Font(bold=True, color=C_WHITE, size=9)
            cell.fill = fill(color)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border()
        ws.row_dimensions[row_num].height = 22

    def auto_width(ws, minimum=12):
        for col in ws.columns:
            max_len = max(
                (len(str(cell.value)) for cell in col if cell.value is not None),
                default=0,
            )
            ws.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 2, minimum)

    today_str = dj_tz.localdate().strftime('%d/%m/%Y')
    product_list = list(products)
    total_products = len(product_list)

    wb = openpyxl.Workbook()

    # ============================================================
    # HOJA 1 — INVENTARIO COMPLETO
    # ============================================================
    ws1 = wb.active
    ws1.title = 'Inventario'

    ws1.merge_cells('A1:N1')
    c = ws1['A1']
    c.value = 'INVENTARIO DE PRODUCTOS — CHE GOLOSO'
    c.font = Font(bold=True, size=14, color=C_WHITE)
    c.fill = fill(C_PURPLE)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[1].height = 28

    ws1.merge_cells('A2:N2')
    c = ws1['A2']
    c.value = f'Generado: {today_str}  |  Total productos activos: {total_products}'
    c.font = Font(size=9, color='555555')
    c.alignment = Alignment(horizontal='center')

    header_row(ws1, 3, [
        'SKU', 'Cód. Barras', 'Nombre', 'Categoría', 'Unidad',
        'P. Compra', 'P. Venta', 'Costo Prom.', 'Margen %',
        'Stock Actual', 'Stock Mín.', 'Stock Máx.',
        'Valor Stock (Costo)', 'Estado',
    ])

    for ri, p in enumerate(product_list, 4):
        margin = float(p.margin_percent)
        if p.current_stock <= 0:
            status, status_fill = 'SIN STOCK', C_RED
        elif p.current_stock <= p.min_stock:
            status, status_fill = 'STOCK BAJO', C_ORANGE
        else:
            status, status_fill = 'Normal', C_GREEN

        bg = C_LGRAY if ri % 2 == 0 else C_WHITE

        row_vals = [
            p.sku,
            p.barcode or '',
            p.name,
            p.category.name if p.category else 'Sin categoría',
            p.unit_of_measure.abbreviation if p.unit_of_measure else 'u',
            float(p.purchase_price),
            float(p.sale_price),
            float(p.cost_price),
            round(margin, 2),
            float(p.current_stock),
            p.min_stock,
            p.max_stock or '',
            float(p.stock_value),
            status,
        ]

        for col, val in enumerate(row_vals, 1):
            cell = ws1.cell(row=ri, column=col, value=val)
            cell.fill = fill(status_fill if col == 14 else bg)
            cell.font = Font(bold=(col == 14), size=9)
            cell.border = border()
            cell.alignment = Alignment(vertical='center')
            if col in (6, 7, 8, 13):
                cell.number_format = '#,##0.00'
            elif col == 9:
                cell.number_format = '0.00'
            elif col == 10:
                cell.number_format = '0.000'

    auto_width(ws1)

    # ============================================================
    # HOJA 2 — STOCK BAJO Y SIN STOCK
    # ============================================================
    ws2 = wb.create_sheet('Stock Bajo y Sin Stock')

    ws2.merge_cells('A1:G1')
    c = ws2['A1']
    c.value = 'ALERTAS DE STOCK — PRODUCTOS CON STOCK BAJO O SIN STOCK'
    c.font = Font(bold=True, size=13, color=C_WHITE)
    c.fill = fill(C_PINK)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 26

    header_row(ws2, 2, [
        'SKU', 'Nombre', 'Categoría', 'Unidad',
        'Stock Actual', 'Stock Mínimo', 'Estado',
    ])

    alert_products = [p for p in product_list if p.current_stock <= p.min_stock]
    if alert_products:
        for ri, p in enumerate(alert_products, 3):
            if p.current_stock <= 0:
                status, bg = 'SIN STOCK', C_RED
            else:
                status, bg = 'STOCK BAJO', C_ORANGE
            for col, val in enumerate([
                p.sku,
                p.name,
                p.category.name if p.category else 'Sin categoría',
                p.unit_of_measure.abbreviation if p.unit_of_measure else 'u',
                float(p.current_stock),
                p.min_stock,
                status,
            ], 1):
                cell = ws2.cell(row=ri, column=col, value=val)
                cell.fill = fill(bg if col == 7 else C_LGRAY)
                cell.font = Font(bold=(col == 7), size=9)
                cell.border = border()
                cell.alignment = Alignment(vertical='center')
    else:
        ws2.cell(row=3, column=1, value='✓ No hay productos con stock bajo').font = Font(
            color='006600', bold=True, size=10
        )

    auto_width(ws2)

    # ============================================================
    # HOJA 3 — ANÁLISIS DE MÁRGENES Y PRECIOS
    # ============================================================
    ws3 = wb.create_sheet('Análisis de Márgenes')

    ws3.merge_cells('A1:G1')
    c = ws3['A1']
    c.value = 'ANÁLISIS DE MÁRGENES Y RENTABILIDAD'
    c.font = Font(bold=True, size=13, color=C_WHITE)
    c.fill = fill(C_PURPLE)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws3.row_dimensions[1].height = 26

    header_row(ws3, 2, [
        'SKU', 'Nombre', 'Categoría',
        'P. Compra', 'P. Venta', 'Ganancia Unit.', 'Margen %',
    ])

    for ri, p in enumerate(sorted(product_list, key=lambda x: -float(x.margin_percent)), 3):
        margin = float(p.margin_percent)
        if margin < 10:
            bg = C_RED
        elif margin < 20:
            bg = C_ORANGE
        else:
            bg = C_LGRAY if ri % 2 == 0 else C_WHITE

        for col, val in enumerate([
            p.sku,
            p.name,
            p.category.name if p.category else 'Sin categoría',
            float(p.purchase_price),
            float(p.sale_price),
            float(p.profit),
            round(margin, 2),
        ], 1):
            cell = ws3.cell(row=ri, column=col, value=val)
            cell.fill = fill(bg)
            cell.font = Font(size=9)
            cell.border = border()
            cell.alignment = Alignment(vertical='center')
            if col in (4, 5, 6):
                cell.number_format = '#,##0.00'
            elif col == 7:
                cell.number_format = '0.00'

    # Leyenda colores
    legend_row = total_products + 4
    for col, (text, color) in enumerate([
        ('Margen < 10% — Revisar precio', C_RED),
        ('Margen 10-20% — Margen bajo',   C_ORANGE),
        ('Margen > 20% — Normal',         C_LGRAY),
    ], 1):
        cell = ws3.cell(row=legend_row, column=col * 2 - 1, value=text)
        cell.fill = fill(color)
        cell.font = Font(size=8, italic=True)
        cell.border = border()

    auto_width(ws3)

    # ============================================================
    # HOJA 4 — VALOR DEL INVENTARIO POR CATEGORÍA
    # ============================================================
    ws4 = wb.create_sheet('Valor de Inventario')

    ws4.merge_cells('A1:F1')
    c = ws4['A1']
    c.value = 'VALOR TOTAL DEL INVENTARIO POR CATEGORÍA'
    c.font = Font(bold=True, size=13, color=C_WHITE)
    c.fill = fill(C_PURPLE)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws4.row_dimensions[1].height = 26

    header_row(ws4, 2, [
        'Categoría', 'N° Productos', 'Valor a Costo', 'Valor a P. Venta', 'Ganancia Potencial', '% del Total',
    ])

    by_cat = {}
    for p in product_list:
        cat = p.category.name if p.category else 'Sin categoría'
        if cat not in by_cat:
            by_cat[cat] = {'cost': 0.0, 'sale': 0.0, 'count': 0}
        by_cat[cat]['cost']  += float(p.stock_value)
        by_cat[cat]['sale']  += float(p.stock_value_sale)
        by_cat[cat]['count'] += 1

    total_cost = sum(d['cost'] for d in by_cat.values())
    total_sale = sum(d['sale'] for d in by_cat.values())

    for ri, (cat_name, d) in enumerate(sorted(by_cat.items()), 3):
        bg = C_LGRAY if ri % 2 == 0 else C_WHITE
        pct = (d['cost'] / total_cost * 100) if total_cost else 0.0
        for col, val in enumerate([
            cat_name, d['count'], d['cost'], d['sale'], d['sale'] - d['cost'], round(pct, 2),
        ], 1):
            cell = ws4.cell(row=ri, column=col, value=val)
            cell.fill = fill(bg)
            cell.font = Font(size=9)
            cell.border = border()
            if col in (3, 4, 5):
                cell.number_format = '#,##0.00'
            elif col == 6:
                cell.number_format = '0.00'

    # Total row
    tr = len(by_cat) + 3
    for col, val in enumerate([
        'TOTAL', total_products, total_cost, total_sale, total_sale - total_cost, 100.0,
    ], 1):
        cell = ws4.cell(row=tr, column=col, value=val)
        cell.font = Font(bold=True, color=C_WHITE)
        cell.fill = fill(C_PINK)
        cell.border = border()
        if col in (3, 4, 5):
            cell.number_format = '#,##0.00'
        elif col == 6:
            cell.number_format = '0.00'

    auto_width(ws4)

    # --- output ---
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="inventario_{dj_tz.localdate().strftime("%Y%m%d")}.xlsx"'
    )
    wb.save(response)
    return response


@login_required
def api_get_packaging(request, packaging_id):
    """API para obtener datos de un empaque por ID (para edición)."""
    try:
        pkg = ProductPackaging.objects.get(pk=packaging_id)
        return JsonResponse({
            'success': True,
            'id': pkg.id,
            'packaging_type': pkg.packaging_type,
            'barcode': pkg.barcode,
            'name': pkg.name,
            'units_per_display': pkg.units_per_display,
            'displays_per_bulk': pkg.displays_per_bulk,
            'units_quantity': pkg.units_quantity,
            'purchase_price': str(pkg.purchase_price),
            'sale_price': str(pkg.sale_price),
            'margin_percent': str(pkg.margin_percent),
            'is_default': pkg.is_default,
            'is_active': pkg.is_active,
        })
    except ProductPackaging.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Empaque no encontrado'}, status=404)


@login_required
@group_required(['Admin', 'Cajero Manager'])
def packaging_delete(request, packaging_id):
    """Eliminar un empaque."""

    packaging = get_object_or_404(ProductPackaging, pk=packaging_id)
    product_id = packaging.product.id
    packaging_name = str(packaging)

    if request.method == 'POST':
        packaging.delete()
        messages.success(request, f'Empaque "{packaging_name}" eliminado.')
        return redirect('stocks:product_packaging', pk=product_id)

    return redirect('stocks:product_packaging', pk=product_id)


# ==================== GESTIÓN DE EMPAQUES ====================

@login_required
@group_required(['Admin', 'Cajero Manager'])
def product_packaging_view(request, pk):
    """Vista completa de gestión de empaques con recepción, apertura y ajuste."""
    product = get_object_or_404(Product, pk=pk)

    unit_pkg = product.packagings.filter(packaging_type='unit', is_active=True).first()
    display_pkg = product.packagings.filter(packaging_type='display', is_active=True).first()
    bulk_pkg = product.packagings.filter(packaging_type='bulk', is_active=True).first()

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'receive':
            pkg_id = request.POST.get('packaging_id')
            qty = request.POST.get('quantity', '0')
            cost = request.POST.get('cost', '').strip()
            try:
                pkg = ProductPackaging.objects.get(pk=pkg_id, product=product)
                StockManagementService.receive_packaging(
                    pkg,
                    Decimal(qty),
                    cost=Decimal(cost) if cost else None,
                    user=request.user,
                )
                messages.success(request, f'Recibido {qty} {pkg.get_packaging_type_display()}(s) de {product.name}.')
            except Exception as e:
                messages.error(request, f'Error al recibir mercadería: {e}')

        elif action == 'adjust':
            pkg_id = request.POST.get('packaging_id')
            new_stock = request.POST.get('new_stock', '0')
            reason = request.POST.get('reason', '')
            try:
                pkg = ProductPackaging.objects.get(pk=pkg_id, product=product)
                old_stock = pkg.current_stock
                new_stock_val = Decimal(new_stock)
                diff = new_stock_val - old_stock

                # Calcular diferencia en unidades base
                if pkg.packaging_type == 'bulk':
                    units_diff = diff * Decimal(str(pkg.units_quantity))
                elif pkg.packaging_type == 'display':
                    units_diff = diff * Decimal(str(pkg.units_per_display))
                else:
                    units_diff = diff

                # Actualizar el packaging ajustado
                pkg.current_stock = new_stock_val
                pkg.save()

                # Actualizar Product.current_stock
                stock_before = product.current_stock
                product.current_stock = stock_before + units_diff
                product.save()

                # Actualizar otros niveles de packaging proporcionalmente
                other_pkgs = product.packagings.filter(
                    is_active=True
                ).exclude(pk=pkg.pk)
                for other in other_pkgs:
                    if other.packaging_type == 'unit':
                        other.current_stock += units_diff
                    elif other.packaging_type == 'display':
                        if other.units_per_display > 0:
                            other.current_stock += units_diff / Decimal(str(other.units_per_display))
                    elif other.packaging_type == 'bulk':
                        if other.units_quantity > 0:
                            other.current_stock += units_diff / Decimal(str(other.units_quantity))
                    other.save()

                StockMovement.objects.create(
                    product=product,
                    movement_type='adjustment_in' if diff >= 0 else 'adjustment_out',
                    quantity=units_diff,
                    stock_before=stock_before,
                    stock_after=product.current_stock,
                    reference=f'Ajuste {pkg.get_packaging_type_display()}',
                    notes=reason,
                    created_by=request.user,
                )
                messages.success(request, f'Stock de {pkg.get_packaging_type_display()} ajustado a {new_stock}.')
            except Exception as e:
                messages.error(request, f'Error al ajustar stock: {e}')

        elif action == 'open':
            pkg_id = request.POST.get('packaging_id')
            qty = request.POST.get('quantity', '1')
            try:
                pkg = ProductPackaging.objects.get(pk=pkg_id, product=product)
                StockManagementService.open_packaging(pkg, Decimal(qty), user=request.user)
                messages.success(request, f'Empaque abierto correctamente.')
            except Exception as e:
                messages.error(request, f'Error al abrir empaque: {e}')

        elif action == 'save_pkg':
            try:
                _save_inline_packaging(request, product)
                # Sync prices with the product base
                bulk_pkg_saved = product.packagings.filter(packaging_type='bulk', is_active=True).first()
                if bulk_pkg_saved:
                    _sync_packaging_prices(product, bulk_pkg_saved)
                messages.success(request, 'Empaques actualizados correctamente.')
            except Exception as e:
                messages.error(request, f'Error al guardar empaques: {e}')

        return redirect('stocks:product_packaging', pk=product.pk)

    # Equivalencias
    equiv_rows = []
    total_equiv = Decimal('0')
    for label, pkg in [('Bultos', bulk_pkg), ('Displays', display_pkg), ('Unidades', unit_pkg)]:
        if pkg:
            equiv = pkg.current_stock * Decimal(str(pkg.units_quantity))
            equiv_rows.append({
                'label': label,
                'stock': pkg.current_stock,
                'equiv': equiv,
                'sale_price': pkg.sale_price,
                'min_stock': pkg.min_stock,
                'type': pkg.packaging_type,
            })
            total_equiv += equiv

    context = {
        'product': product,
        'unit_pkg': unit_pkg,
        'display_pkg': display_pkg,
        'bulk_pkg': bulk_pkg,
        'packaging_cards': [
            ('Bulto', bulk_pkg, 'bulk', 'fa-cubes'),
            ('Display', display_pkg, 'display', 'fa-box'),
            ('Unidad', unit_pkg, 'unit', 'fa-cube'),
        ],
        'equiv_rows': equiv_rows,
        'total_equiv': total_equiv,
        'form': ProductPackagingForm(),
        'categories': ProductCategory.objects.filter(is_active=True),
    }
    return render(request, 'stocks/product_packaging.html', context)


@login_required
@group_required(['Admin', 'Cajero Manager'])
def packaging_inventory_view(request):
    """Inventario general de empaques."""
    qs = Product.objects.filter(is_active=True).select_related('category').prefetch_related('packagings')

    category_id = request.GET.get('category')
    if category_id:
        qs = qs.filter(category_id=category_id)

    has_packaging = request.GET.get('has_packaging')
    if has_packaging:
        qs = qs.filter(packagings__is_active=True).distinct()

    low_stock = request.GET.get('low_stock')

    rows = []
    for product in qs:
        pkgs = {p.packaging_type: p for p in product.packagings.filter(is_active=True)}
        if has_packaging and not pkgs:
            continue

        unit_pkg = pkgs.get('unit')
        display_pkg = pkgs.get('display')
        bulk_pkg = pkgs.get('bulk')

        total_equiv = Decimal('0')
        alert = False
        for pkg in pkgs.values():
            total_equiv += pkg.current_stock * Decimal(str(pkg.units_quantity))
            if pkg.current_stock <= pkg.min_stock:
                alert = True

        if low_stock and not alert:
            continue

        rows.append({
            'product': product,
            'unit_stock': unit_pkg.current_stock if unit_pkg else None,
            'unit_alert': unit_pkg and unit_pkg.current_stock <= unit_pkg.min_stock if unit_pkg else False,
            'display_stock': display_pkg.current_stock if display_pkg else None,
            'display_alert': display_pkg and display_pkg.current_stock <= display_pkg.min_stock if display_pkg else False,
            'bulk_stock': bulk_pkg.current_stock if bulk_pkg else None,
            'bulk_alert': bulk_pkg and bulk_pkg.current_stock <= bulk_pkg.min_stock if bulk_pkg else False,
            'total_equiv': total_equiv,
            'has_alert': alert,
        })

    paginator = Paginator(rows, 25)
    page = paginator.get_page(request.GET.get('page'))

    context = {
        'page_obj': page,
        'categories': ProductCategory.objects.filter(is_active=True),
        'current_category': category_id,
        'current_low_stock': low_stock,
        'current_has_packaging': has_packaging,
    }
    return render(request, 'stocks/packaging_inventory.html', context)


@login_required
def packaging_api(request, pk):
    """API JSON con los packagings de un producto."""
    product = get_object_or_404(Product, pk=pk)
    data = []
    for pkg in product.packagings.filter(is_active=True):
        data.append({
            'id': pkg.id,
            'packaging_type': pkg.packaging_type,
            'type_display': pkg.get_packaging_type_display(),
            'barcode': pkg.barcode,
            'name': pkg.name,
            'units_quantity': pkg.units_quantity,
            'units_per_display': pkg.units_per_display,
            'displays_per_bulk': pkg.displays_per_bulk,
            'purchase_price': str(pkg.purchase_price),
            'sale_price': str(pkg.sale_price),
            'margin_percent': str(pkg.margin_percent),
            'current_stock': str(pkg.current_stock),
            'min_stock': pkg.min_stock,
        })
    return JsonResponse({'success': True, 'packagings': data})
